import random
import openpyxl

n = int(input('Введите количество вершин в графе:'))
edge_probability = float(input('Введите вероятность существования ребра между вершинами : '))
p = int(n * (n - 1) / 2 * edge_probability)  # Вычисляем количество рёбер по вероятности

edges = set()

while len(edges) < p:
    random_numbers = [random.randint(0, n - 1), random.randint(0, n - 1)]

    # Уникальность исходя из трех условий
    is_unique = True
    for edge in edges:
        if random_numbers[0] == random_numbers[1] or ' '.join(map(str, random_numbers)) in edges or ' '.join(map(str, random_numbers[::-1])) in edges:
            is_unique = False
            break

    if is_unique:
        edges.add(' '.join(map(str, random_numbers)))

print('Создан граф из',n,'вершин и',len(edges),'ребер. Рёбра графа:')
# Вывод уникальных ребер
for edge in edges:
    print(edge)


# Создаем таблицу Excel
wb = openpyxl.Workbook()
sheet = wb.active


# Записываем данные в таблицу
for i, line in enumerate(edges):
    num1, num2 = line.split()
    sheet.cell(row=i+1, column=1, value=int(num1)) #строка, column-столбец, значение
    sheet.cell(row=i+1, column=2, value=int(num2))

# Сохраняем книгу
wb.save('edges.xlsx')
print('данные сохранены')
probability_of_edge=edge_probability
